import os
import json
from openpyxl import load_workbook
from flask import Flask, request
from flask_restful import reqparse, abort, Resource, Api
from flasgger import Swagger

app = Flask(__name__)
api = Api(app)
swagger = Swagger(app)
path_dir = os.getcwd()

def abort_if_param_doesnt_exist(dict_from_json, id, type_field):
    if id not in dict_from_json.keys():
        abort(404, message="{} doesn't exist".format(id))
    if type_field not in dict_from_json.values():
        abort(404, message="{} doesn't exist".format(type_field))

def read_json(id, type_field):
    with open(path_dir + '/data/тест.json') as f:
        invalid_json = f.read()
    valid_json = invalid_json.replace("'", '"')
    json.loads(valid_json)

    with open(path_dir + '/data/тест.json', 'w') as f:
        f.write(valid_json)
    
    with open(path_dir + '/data/тест.json', 'r', encoding='utf-8') as f:
        data_from_json = json.load(f)
    
    dict_from_json = {}
    for key, value in data_from_json.items():
        if key == id or value == type_field:
            dict_from_json[key] = value
    return dict_from_json

def adding_data_from_xlsx(dict_from_json):
    wb = load_workbook(path_dir + '/data/тест.xlsx')
    name_sheet = wb.sheetnames[0]
    ws = wb[name_sheet]
    results_list = []
    for row in ws.rows:
        if row[0].value == 'id':
            continue
        set_res = []
        id_field = row[0].value
        if id_field in dict_from_json.keys():
            type_field = dict_from_json[id_field]
            name = row[1].value
            data = row[2].value
            set_res.append(id_field)
            set_res.append(type_field)
            set_res.append(name)
            set_res.append(data)
            results_list.append(set_res)
    wb.close()
    return results_list


parser = reqparse.RequestParser()
parser.add_argument('id', type=str)
parser.add_argument('name', type=str)


class GetData(Resource):
    def get(self, id, type_field):
        """
       This examples uses FlaskRESTful Resource
       It works also with swag_from, schemas and spec_dict
       ---
       parameters:
         - in: path
           name: username
           type: string
           required: true
       responses:
         200:
           description: A single user item
           schema:
             id: test
             properties:
                id:
                    type: string
                    description: Id
                    default: 1n
                type:
                    type: string
                    description: Type
                    default: тип5
        """
        dict_from_json = read_json(id, type_field)
        abort_if_param_doesnt_exist(dict_from_json, id, type_field)
        results_list = adding_data_from_xlsx(dict_from_json)
        return results_list


class PutData(Resource):
    def post(self):
        request.get_json(force=True)
        args = parser.parse_args()
        id_input = args['id']
        name_input = args['name']
        wb = load_workbook(path_dir + '/data/тест.xlsx')
        name_sheet = wb.sheetnames[0]
        ws = wb[name_sheet]
        n = 0
        for row in ws.rows:
            if row[0].value == id_input:
                row_num = ws._current_row
                cell = 'B' + str(row_num)
                ws[cell] = name_input
                n +=1
        if n == 0:
            abort(404, message="{} doesn't exist".format(id_input))
        wb.save(path_dir + '/data/тест.xlsx')
        response = json.dumps({"info": "ok"})
        return response


api.add_resource(GetData, '/test/service/<string:id>/<string:type_field>')
api.add_resource(PutData, '/test/service')

if __name__ == '__main__':
    app.run(debug=True)